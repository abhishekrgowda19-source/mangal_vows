from flask import Blueprint, render_template, request, redirect, url_for, session, send_from_directory
from models import User, Agent
from database import db
import os

agent_bp = Blueprint("agent_bp", __name__)


def normalize_phone(phone):
    phone = str(phone) if phone else ""
    return ''.join(filter(str.isdigit, phone))[-10:] if phone else ""


@agent_bp.route("/agent-login", methods=["GET", "POST"])
def agent_login():
    if request.method == "POST":
        email    = request.form.get("email", "").strip()
        password = request.form.get("password", "").strip()
        agent    = Agent.query.filter_by(email=email).first()
        from app import bcrypt
        if not agent or not bcrypt.check_password_hash(agent.password_hash, password):
            return render_template("login.html", error="Invalid agent credentials")
        session["agent"] = agent.id
        session["role"]  = "agent"
        return redirect(url_for("agent_bp.agent_dashboard"))
    return render_template("login.html")


@agent_bp.route("/agent-dashboard")
def agent_dashboard():
    if session.get("role") != "agent":
        return redirect(url_for("agent_bp.agent_login"))
    agent = Agent.query.get(session.get("agent"))
    if not agent:
        return redirect(url_for("agent_bp.agent_login"))
    users = User.query.filter_by(agent_id=agent.id).all()
    return render_template("agent_dashboard.html", agent=agent, users=users)


@agent_bp.route("/agent-add-user", methods=["GET", "POST"])
def agent_add_user():
    if session.get("role") != "agent":
        return redirect(url_for("agent_bp.agent_login"))
    agent = Agent.query.get(session.get("agent"))

    if request.method == "POST":
        name          = request.form.get("name", "").strip()
        phone         = normalize_phone(request.form.get("phone", ""))
        email         = request.form.get("email", "").strip()
        age           = request.form.get("age")
        gender        = request.form.get("gender", "").strip()
        height        = request.form.get("height", "").strip()
        religion      = request.form.get("religion", "").strip()
        caste         = request.form.get("caste", "").strip()
        community     = request.form.get("community", "").strip()
        mother_tongue = request.form.get("mother_tongue", "").strip()
        profession    = request.form.get("profession", "").strip()
        education     = request.form.get("education", "").strip()
        city          = request.form.get("city", "").strip()
        state         = request.form.get("state", "").strip()

        if not name or not phone:
            return render_template("register.html", error="Name and phone are required")
        if User.query.filter_by(phone=phone).first():
            return render_template("register.html", error="Phone already registered")

        new_user = User(
            name=name, phone=phone, email=email,
            age=int(age) if age else None,
            gender=gender, height=height, religion=religion,
            caste=caste, community=community, mother_tongue=mother_tongue,
            profession=profession, education=education,
            city=city, state=state,
            location=f"{city}, {state}".strip(", "),
            agent_id=agent.id
        )
        db.session.add(new_user)
        db.session.commit()
        return redirect(url_for("agent_bp.agent_dashboard"))

    return render_template("register.html")


# ── Bulk Upload ───────────────────────────────────────

@agent_bp.route("/agent-bulk-upload", methods=["GET", "POST"])
def agent_bulk_upload():
    if session.get("role") != "agent":
        return redirect(url_for("agent_bp.agent_login"))

    agent = Agent.query.get(session.get("agent"))

    if request.method == "POST":
        file = request.files.get("excel_file")

        if not file or file.filename == "":
            return render_template("bulk_upload.html", agent=agent,
                                   error="Please select an Excel file")

        if not file.filename.endswith((".xlsx", ".xls")):
            return render_template("bulk_upload.html", agent=agent,
                                   error="Only .xlsx or .xls files allowed")

        try:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            # ✅ Read headers from row 3
            headers = []
            for cell in ws[3]:
                headers.append(cell.value)

            success_count = 0
            skip_count    = 0
            errors        = []

            # ✅ Data starts from row 5 (row 4 is notes, rows 5-7 are samples)
            for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):

                # Skip empty rows
                if not any(row):
                    continue

                # Skip sample/example rows
                row_data = dict(zip(headers, row))

                name  = str(row_data.get("name", "") or "").strip()
                phone = normalize_phone(str(row_data.get("phone", "") or ""))

                # Skip if name or phone missing
                if not name or not phone or len(phone) < 10:
                    skip_count += 1
                    continue

                # Skip if phone already exists
                if User.query.filter_by(phone=phone).first():
                    skip_count += 1
                    errors.append(f"Row {row_idx}: {name} ({phone}) already exists")
                    continue

                age = row_data.get("age")
                try:
                    age = int(age) if age else None
                except:
                    age = None

                city  = str(row_data.get("city",  "") or "").strip()
                state = str(row_data.get("state", "") or "").strip()

                new_user = User(
                    name          = name,
                    phone         = phone,
                    email         = str(row_data.get("email", "")         or "").strip(),
                    age           = age,
                    gender        = str(row_data.get("gender", "")        or "").strip(),
                    height        = str(row_data.get("height", "")        or "").strip(),
                    religion      = str(row_data.get("religion", "")      or "").strip(),
                    caste         = str(row_data.get("caste", "")         or "").strip(),
                    community     = str(row_data.get("community", "")     or "").strip(),
                    mother_tongue = str(row_data.get("mother_tongue", "") or "").strip(),
                    profession    = str(row_data.get("profession", "")    or "").strip(),
                    education     = str(row_data.get("education", "")     or "").strip(),
                    city          = city,
                    state         = state,
                    location      = f"{city}, {state}".strip(", "),
                    agent_id      = agent.id
                )
                db.session.add(new_user)
                success_count += 1

            db.session.commit()

            return render_template(
                "bulk_upload.html",
                agent=agent,
                success=True,
                success_count=success_count,
                skip_count=skip_count,
                errors=errors
            )

        except Exception as e:
            return render_template("bulk_upload.html", agent=agent,
                                   error=f"Error reading file: {str(e)}")

    return render_template("bulk_upload.html", agent=agent)


# ── Download Template ─────────────────────────────────

@agent_bp.route("/download-template")
def download_template():
    if session.get("role") != "agent":
        return redirect(url_for("agent_bp.agent_login"))
    static_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "static")
    return send_from_directory(
        static_dir,
        "mangal_vows_upload_template.xlsx",
        as_attachment=True
    )


@agent_bp.route("/agent-logout")
def agent_logout():
    session.clear()
    return redirect(url_for("user.home"))